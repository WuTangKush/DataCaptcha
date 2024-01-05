from tkinter import *
from tkinter import messagebox
from datetime import datetime
import os
import openpyxl

file_name = "data.xlsx"

try:

    if os.path.isfile(file_name):
        workbook = openpyxl.load_workbook(file_name)
    else:
        workbook = openpyxl.Workbook()
        workbook.save(file_name)
except Exception as e:
    print(f"Error access workbook: {e}")

sheet = workbook.active

def submit_form(sheet, name, selected_order, delivery_mode, location, payment_mode, amount_billed):
    
    current_time = datetime.now()
    current_date = current_time.date()
    print("Form submitted on:", current_date)
    print("Name:", name)
    print("Selected Order:", selected_order)
    print("Delivery Mode:", delivery_mode)
    print("Location:", location)
    print("Payment Mode:", payment_mode)
    print("Amount Billed:", amount_billed)

    #Write data to Excel Sheet
    next_row = sheet.max_row + 1
    sheet.cell(row = next_row, column = 1, value = current_date)
    sheet.column_dimensions['A'].width = 15
    sheet.cell(row=next_row, column=1).number_format = 'dd-mm-yyyy'
    sheet.cell(row = next_row, column = 2, value = name)
    sheet.cell(row = next_row, column = 3, value = selected_order)
    sheet.cell(row = next_row, column = 4, value = delivery_mode)
    sheet.cell(row = next_row, column = 5, value = location)
    sheet.cell(row = next_row, column = 6, value = amount_billed)
    
    workbook.save(file_name)

    name_entry.delete(0, END)
    location_entry.delete(0, END)
    amount_entry.delete(0, END)
    order_option.set(options[0])
   

    messagebox.showinfo("Entry added")

# Main window
main = Tk()
main.title("DPW")

# Label
info_label = Label(main, text="Enter sale details", font=("Helvetica", 12))
info_label.pack()

# Frame
details_Frame = Frame(main, width=200, height=200, padx=20, pady=20)

# Frame labels
name_label = Label(details_Frame, text="Name")
order_label = Label(details_Frame, text="Order")
delivery_label = Label(details_Frame, text="Delivery Mode")
location_label = Label(details_Frame, text="Location")
amount_label = Label(details_Frame, text="Amount Billed")

# Frame labels positions
name_label.grid(row=0, column=0)
order_label.grid(row=1, column=0)
delivery_label.grid(row=2, column=0, columnspan=2)
location_label.grid(row=6, column=0)
amount_label.grid(row=8, column=0)

# Frame entries
name_entry = Entry(details_Frame)
location_entry = Entry(details_Frame)
amount_entry = Entry(details_Frame)

# Frame entries positions
name_entry.grid(row=0, column=1)
location_entry.grid(row=6, column=1)
amount_entry.grid(row=8, column=1)

# Frame dropdown menu
order_option = StringVar(main)
order_option.set("Carbonara")
options = ["Carbonara", "Fillet", "Fish Fingers"]
option_menu = OptionMenu(details_Frame, order_option, *options)
option_menu.grid(row=1, column=1)

# Frame radio buttons
delivery_mode = StringVar()
radio1 = Radiobutton(details_Frame, text="Walk In", variable=delivery_mode, value="Walk In")
radio2 = Radiobutton(details_Frame, text="Glovo", variable=delivery_mode, value="Glovo")
radio3 = Radiobutton(details_Frame, text="Jumia", variable=delivery_mode, value="Jumia")
radio4 = Radiobutton(details_Frame, text="Soulful Delivery", variable=delivery_mode, value="Delivery")

payment_mode = StringVar()
cash_radio = Radiobutton(details_Frame, text="Cash", variable=payment_mode, value="Cash")
mpesa_radio = Radiobutton(details_Frame, text="Mpesa", variable=payment_mode, value="Mpesa")

# Frame radio buttons positions
radio1.grid(row=3, column=0)
radio2.grid(row=3, column=1)
radio3.grid(row=4, column=0)
radio4.grid(row=4, column=1)

cash_radio.grid(row=7, column=0)
mpesa_radio.grid(row=7, column=1)

# Save Button
save_button = Button(details_Frame, text="Save", command=lambda: submit_form(
    sheet,
    name_entry.get(),
    order_option.get(),
    delivery_mode.get(),
    location_entry.get(),
    payment_mode.get(),
    amount_entry.get()
))
save_button.grid(row=9, column=0, columnspan=2)

details_Frame.pack()

main.mainloop()
